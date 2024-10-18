import openpyxl
import logging

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


def get_row_count(file, sheetname):
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetname]
        row_count = sheet.max_row
        logger.info(f"Total number of rows in '{sheetname}': {row_count}")
        return row_count
    except Exception as e:
        logger.error(f"Error in getting row count: {e}")
        return None


def get_column_count(file, sheetname):
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetname]
        col_count = sheet.max_column
        logger.info(f"Total number of columns in '{sheetname}': {col_count}")
        return col_count
    except Exception as e:
        logger.error(f"Error in getting column count: {e}")
        return None


def read_data(file, sheetname, rownum, columnno):
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetname]
        cell_value = sheet.cell(row=rownum, column=columnno).value
        logger.info(f"Read data from {sheetname} - Row: {rownum}, Column: {columnno} - Value: {cell_value}")
        return cell_value
    except Exception as e:
        logger.error(f"Error in reading data: {e}")
        return None


def write_data(file, sheetname, rownum, columnno, data):
    try:
        workbook = openpyxl.load_workbook(file)
        sheet = workbook[sheetname]
        sheet.cell(row=rownum, column=columnno).value = data
        workbook.save(file)
        logger.info(f"Written data to {sheetname} - Row: {rownum}, Column: {columnno} - Data: {data}")
    except Exception as e:
        logger.error(f"Error in writing data: {e}")
